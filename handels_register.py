import argparse
import mechanize
import pathlib
import sys
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
import zipfile
import requests
import os
import pandas as pd
from tqdm import tqdm
import xml.etree.ElementTree as ET
from lxml import etree
import concurrent.futures

# Dictionaries to map arguments to values
schlagwortOptionen = {
    "all": 1,
    "min": 2,
    "exact": 3
}

class XMLParser:
    def __init__(self, xml_file_path):
        self.xml_file_path = xml_file_path
        self.tree = None
        self.root = None

    def parse_xml(self):
        try:
            parser = etree.XMLParser(recover=True)
            self.tree = etree.parse(self.xml_file_path, parser)
            self.root = self.tree.getroot()
            print("XML parsed successfully.")
        except etree.XMLSyntaxError as e:
            print(f"Error parsing XML: {e}")
            raise
        except FileNotFoundError as e:
            print(f"File not found: {e}")
            raise
    
    def get_element_text(self, element_path, namespaces=None):
        element = self.root.find(element_path, namespaces=namespaces)
        return element.text.strip() if element is not None and element.text else None
    
    def get_comment_from_element(self, element_path, namespaces=None):
        element = self.root.xpath(element_path, namespaces=namespaces)
        if element:
            for comment in self.root.xpath('//comment()'):
                if comment.getparent() == element[0]:
                    return comment.text.strip()
        return None

    def retrieve_xml_data(self, namespaces):
        """
        This function will parse and retrieve elements from the XML.
        """
        results = []

        # Find all <tns:vollerName> elements
        vollerName_elements = self.root.findall('.//tns:vollerName', namespaces=namespaces)

        # Count the occurrences of <tns:vollerName>
        count_vollerName = len(vollerName_elements)
    
        for i in range(count_vollerName):
            bezeichnung_aktuell = self.get_element_text('.//tns:bezeichnung.aktuell', namespaces=namespaces)
            anschrift_strasse = self.get_element_text('.//tns:anschrift/tns:strasse', namespaces=namespaces)
            anschrift_hausnummer = self.get_element_text('.//tns:anschrift/tns:hausnummer', namespaces=namespaces)
            anschrift_postleitzahl = self.get_element_text('.//tns:anschrift/tns:postleitzahl', namespaces=namespaces)
            anschrift_ort = self.get_element_text('.//tns:anschrift/tns:ort', namespaces=namespaces)

            vorname = vollerName_elements[i].find('tns:vorname', namespaces=namespaces).text
            nachname = vollerName_elements[i].find('tns:nachname', namespaces=namespaces).text
            geburtsdatum = self.root.findall('.//tns:geburtsdatum', namespaces=namespaces)[i].text  

            geschlecht = self.get_comment_from_element('.//tns:geschlecht', namespaces=namespaces)
            rechtsform_comment = self.get_comment_from_element('.//tns:angabenZurRechtsform/tns:rechtsform', namespaces=namespaces)
            gegenstand = self.get_element_text('.//tns:basisdatenRegister/tns:gegenstand', namespaces=namespaces)
            vertretungsbefugnis = self.get_element_text('.//tns:auswahl_vertretungsbefugnis/tns:vertretungsbefugnisFreitext', namespaces=namespaces)
            
            result = {
                "bezeichnung":bezeichnung_aktuell,
                "rechtsform":rechtsform_comment,
                "strasse":anschrift_strasse,
                "hausnummer":anschrift_hausnummer,
                "postleitzahl":anschrift_postleitzahl,
                "ort":anschrift_ort,
                "vorname":vorname,
                "nachname":nachname,
                "geschlecht":geschlecht,
                "geburtsdatum":geburtsdatum,
                "gegenstand":gegenstand,
                "vertretungsbefugnis":vertretungsbefugnis
            }

            results.append(result)

        return results


class HandelsRegister:
    def __init__(self, args):
        self.args = args
        self.xml_parser = None  # init xml_parser
        self.browser = mechanize.Browser()

        self.browser.set_debug_http(args.debug)
        self.browser.set_debug_responses(args.debug)

        self.browser.set_handle_robots(False)
        self.browser.set_handle_equiv(True)
        self.browser.set_handle_gzip(True)
        self.browser.set_handle_refresh(False)
        self.browser.set_handle_redirect(True)
        self.browser.set_handle_referer(True)

        self.browser.addheaders = [
            (
                "User-Agent",
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/15.5 Safari/605.1.15",
            ),
            ("Accept-Language", "en-GB,en;q=0.9"),
            ("Accept-Encoding", "gzip, deflate, br"),
            (
                "Accept",
                "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            ),
            ("Connection", "keep-alive"),
        ]

        self.cachedir = pathlib.Path("cache")
        self.cachedir.mkdir(parents=True, exist_ok=True)

    def open_startpage(self):
        self.browser.open("https://www.handelsregister.de/rp_web/welcome.xhtml", timeout=10)

    def companyname2cachename(self, companyname):
        return self.cachedir / companyname

    def search_company(self):
        cachename = self.companyname2cachename(self.args.schlagwoerter)
        cookie_dict = {}  # Initialize cookie_dict 

        # Check if the cache file exists
        if cachename.exists():
            try:
                # Delete the existing cache file
                os.remove(cachename)
                print(f"Deleted existing cache file: {cachename}")
            except Exception as e:
                print(f"Error while deleting file {cachename}: {e}")

        if self.args.force == False and cachename.exists():
            with open(cachename, "r") as f:
                html = f.read()
                print("return cached content for %s" % self.args.schlagwoerter)
        else:
            response = self.browser.open("https://www.handelsregister.de/rp_web/erweitertesuche.xhtml")
            if self.args.debug:
                print(self.browser.title())

            self.browser.select_form(name="form")

            self.browser["form:schlagwoerter"] = self.args.schlagwoerter
            so_id = schlagwortOptionen.get(self.args.schlagwortOptionen)

            self.browser["form:schlagwortOptionen"] = [str(so_id)]

            response_result = self.browser.submit()

            if self.args.debug:
                print(self.browser.title())

            html = response_result.read().decode("utf-8")

            
            # with open(cachename, "w") as f:
            #     f.write(html)
            
            # Capture cookies from the mechanize browser
            cookies = self.browser._ua_handlers['_cookies'].cookiejar
            cookie_dict = {cookie.name: cookie.value for cookie in cookies}

        return html, cookie_dict

    def get_companies_xml_file(self, soup, cookies, company_name):
        # Find the table body containing the results
        tbody = soup.find('tbody', id='ergebnissForm:selectedSuchErgebnisFormTable_data')

        if not tbody:
            print("No results table found. This error maybe probably due to incorrect company name.")
            return False # Exit the function if the tbody is not found

        # Find the first row within the results table
        first_row = tbody.find('tr')

        if not first_row:
            print("No rows found in the results table. This error maybe probably due to incorrect company name.")
            return False # Exit the function if no rows are found
            
        # Find all <a> tags within the first row
        a_tags = first_row.find_all('a')

        # Check if any <a> tags are found
        if not a_tags:
            print("No <a> tags found in the first row. This error maybe probably due to incorrect company name.")
            return False # Exit the function if no <a> tags are found

        # Get the last <a> tag
        last_a_tag = first_row.find_all('a')[-1]
        last_a_id = last_a_tag.get('id')

        # Find the form element by ID
        form_element = soup.find('form', id='ergebnissForm')
        action_url = form_element['action']
        query_string = action_url.split('?')[1] if '?' in action_url else ''

        # Get the ViewState value
        view_state_element = soup.find('input', {'name': 'javax.faces.ViewState'})
        view_state = view_state_element.get('value') if view_state_element else None

        # Get the JSESSIONID cookie value
        jsessionid_cookie = cookies.get("JSESSIONID", "")

        # Construct the URL for the POST request
        url = f"https://www.handelsregister.de/rp_web/xhtml/research/sucheErgebnisse.xhtml?{query_string}"

        headers = {
            "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
            "accept-encoding": "gzip, deflate, br, zstd",
            "accept-language": "en-GB,en-US;q=0.9,en;q=0.8",
            "cache-control": "max-age=0",
            "connection": "keep-alive",
            "content-type": "application/x-www-form-urlencoded",
            "cookie": f"JSESSIONID={jsessionid_cookie}",
            "host": "www.handelsregister.de",
            "origin": "https://www.handelsregister.de",
            "referer": "https://www.handelsregister.de/",
            "sec-ch-ua": '"Not)A;Brand";v="99", "Google Chrome";v="127", "Chromium";v="127"',
            "sec-ch-ua-mobile": "?0",
            "sec-ch-ua-platform": '"Windows"',
            "sec-fetch-dest": "document",
            "sec-fetch-mode": "navigate",
            "sec-fetch-site": "same-origin",
            "sec-fetch-user": "?1",
            "upgrade-insecure-requests": "1",
            "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/127.0.0.0 Safari/537.36"
        }

        data = {
            "ergebnissForm": "ergebnissForm",
            "ergebnissForm:selectedSuchErgebnisFormTable_rppDD": "10",
            "javax.faces.ViewState": f"{view_state}",
            "property2": "",
            f"{last_a_id}": f"{last_a_id}",
            "property": "Global.Dokumentart.SI"
        }

        response = requests.post(url, headers=headers, data=data)
        company_name = str(company_name).replace('/', '')

        # Save the XML file
        with open(f"files/{company_name}.xml", "wb") as file:
            file.write(response.content)

        print("File downloaded successfully")

    def get_companies_in_searchresults(self, html, cookies, xml_file_path, company_name):
        soup = BeautifulSoup(html, 'html.parser')
        grid = soup.find('table', role='grid')

        # Call the new function to download the XML file
        self.get_companies_xml_file(soup, cookies, company_name=company_name)

        # # Initialize XMLParser
        # self.xml_parser = XMLParser(xml_file_path)
        #
        # # Parse the XML
        # try:
        #     self.xml_parser.parse_xml()
        # except Exception as e:
        #     print(f"Error parsing XML file: {e}")
        #     return []
        #
        # # Define namespace
        # ns = {'tns': 'http://www.xjustiz.de'}
        #
        # # Retrieve XML data
        # xml_data = self.xml_parser.retrieve_xml_data(ns)
        #
        # results_i = []
        # merged_data_i = []
        #
        # for x in range(len(xml_data)):
        #     results = []
        #     for result in grid.find_all('tr'):
        #         a = result.get('data-ri')
        #         if a is not None:
        #             d = self.parse_result(result)
        #             results.append(d)
        #     results_i.append(results)
        #
        #     # Merge results and xml_data[x]
        #     merged_data = []
        #     for i in range(len(results)):
        #         merged_entry = {**results[i], **xml_data[x]}
        #         merged_entry.pop('documents', None)
        #         merged_entry.pop('history', None)
        #         merged_data.append(merged_entry)
        #     merged_data_i.append(merged_data)
        #
        # return results_i, merged_data_i

    def parse_result(self, result):
        cells = []
        for cellnum, cell in enumerate(result.find_all('td')):
            cells.append(cell.text.strip())

        d = {
            'court': cells[1],  # Gericht
            'name': cells[2],  # Firmenname
            'state': cells[3],  # Sitz
            'status': cells[4],  # Status
            'documents': cells[5],  # Dokumente
            'history': [6]  # Verlauf
        }

        # Extract history if available
        history_cells = result.find_all('td')[8:]
        if history_cells:
            for i in range(0, len(history_cells), 2):
                event = history_cells[i].text.strip()
                date = history_cells[i + 1].text.strip()
                d['history'].append((event, date))

        return d

def save_to_excel(companies, merged_data, filepath):
    try:
        # Attempt to load the existing workbook
        if os.path.exists(filepath):
            workbook = openpyxl.load_workbook(filepath)
        else:
            raise FileNotFoundError
        
        # Access the active sheet or create necessary sheets
        sheet = workbook.active
        sheet_2 = workbook["Goal output"] if "Goal output" in workbook.sheetnames else workbook.create_sheet(title="Goal output")
    
    except FileNotFoundError:
        # Create a new workbook if the file doesn't exist
        workbook = openpyxl.Workbook()

        # Create new sheets and set titles
        sheet = workbook.active
        sheet_2 = workbook.create_sheet(title="Goal output")
        sheet.title = "Current output"
        sheet_2.title = "Goal output"

        sheet.append(["Firmenname", "Gericht", "Sitz", "Status", "Handelsregister-Nummer", "Dokumente", "Verlauf"])
        sheet_2.append(["Company Name", "Court", "City", "Status", "Bezeichnung", "Rechtsform", "Straße", "Hausnummer", "Postleitzahl", "Ort", "Vorname", "Nachname", "Geschlecht", "Geburtsdatum", "Gegenstand", "Vertretungsbefugnis"])

    except (InvalidFileException, zipfile.BadZipFile) as e:
        print(f"Error: The file '{filepath}' is not a valid Excel file or is corrupted: {e}")
        return
    except Exception as e:
        print(f"An unexpected error occurred while loading the Excel file: {e}")
        return


    for company in companies:
        sheet.append([
            company.get("name", ""),
            company.get("court", ""),
            company.get("state", ""),
            company.get("status", ""),
            company.get("documents", ""),
        ])

    # Add or update rows in "Goal output" sheet based on the merged_data
    for company in merged_data:
        name = company.get("name", "")
        court = company.get("court", "")
        state = company.get("state", "")
        status = company.get("status", "")
        bezeichnung = company.get("bezeichnung", "")
        rechtsform = company.get("rechtsform", "")
        strasse = company.get("strasse", "")
        hausnummer = company.get("hausnummer", "")
        postleitzahl = company.get("postleitzahl", "")
        ort = company.get("ort", "")
        vorname = company.get("vorname", "")
        nachname = company.get("nachname", "")
        geschlecht = company.get("geschlecht", "")
        geburtsdatum = company.get("geburtsdatum", "")
        gegenstand = company.get("gegenstand", "")
        vertretungsbefugnis = company.get("vertretungsbefugnis", "")

        # Check if the company name already exists in the "Goal output" sheet
        existing_row = None
        for row in sheet_2.iter_rows(min_row=2, values_only=False):
            if row[0].value == name and row[10].value == vorname:  # Assuming the company name is in the first column
                existing_row = row
                break

        # If the company name exists, update the existing row
        if existing_row:
            existing_row[1].value = court
            existing_row[2].value = state
            existing_row[3].value = status
            existing_row[4].value = bezeichnung
            existing_row[5].value = rechtsform
            existing_row[6].value = strasse
            existing_row[7].value = hausnummer
            existing_row[8].value = postleitzahl
            existing_row[9].value = ort
            existing_row[10].value = vorname
            existing_row[11].value = nachname
            existing_row[12].value = geschlecht
            existing_row[13].value = geburtsdatum
            existing_row[14].value = gegenstand
            existing_row[15].value = vertretungsbefugnis
        else:
            # If the company name does not exist, add a new row
            sheet_2.append([
                name, court, state, status, bezeichnung, rechtsform, strasse,
                hausnummer, postleitzahl, ort, vorname, nachname, geschlecht,
                geburtsdatum, gegenstand, vertretungsbefugnis
            ])

    # Save workbook to Excel file
    try:
        workbook.save(filepath)
        print(f"Data saved to {filepath}")
    except Exception as e:
        print(f"An error occurred while saving the Excel file: {e}")

def parse_args(default_schlagwoerter):
    parser = argparse.ArgumentParser(description='A handelsregister CLI')
    parser.add_argument(
        "-d", 
        "--debug", 
        help="Enable debug mode and activate logging", 
        action="store_true"
    )
    parser.add_argument(
        "-f", 
        "--force", 
        help="Force a fresh pull and skip the cache", 
        action="store_true"
    )
    parser.add_argument(
        "-s", 
        "--schlagwoerter", 
        help="Search for the provided keywords", 
        default=default_schlagwoerter
    )
    parser.add_argument(
        "-so", 
        "--schlagwortOptionen", 
        help="Keyword options: all=contain all keywords; min=contain at least one keyword; exact=contain the exact company name.", 
        choices=["all", "min", "exact"], 
        default="exact"
    )
    parser.add_argument(
        "-o", 
        "--output", 
        help="Path to the output Excel file", 
        default="handelsregister_result.xlsx"
    )
    args = parser.parse_args()

    if args.debug:
        import logging
        logger = logging.getLogger("mechanize")
        logger.addHandler(logging.StreamHandler(sys.stdout))
        logger.setLevel(logging.DEBUG)

    return args

def process_company(company_name, xml_file_path):
    # Get arguments with the current company name as the default schlagwoerter
    args = parse_args(default_schlagwoerter=company_name)

    h = HandelsRegister(args)
    h.open_startpage()
    html, cookies = h.search_company()
    companies = h.get_companies_in_searchresults(html, cookies, xml_file_path, company_name)

    # Ensure there are at least two arrays lists before proceeding
    if companies is not None:
        if len(companies) < 2:
            print(f"Insufficient data for company: {company_name}")
            return

        # Process each company data
        for j in range(min(len(companies[0]), len(companies[1]))):
            try:
                save_to_excel(companies[0][j], companies[1][j], args.output)
                print(f"Ergebnisse wurden in der Datei {args.output} gespeichert.")
            except IndexError as e:
                print(f"IndexError encountered while processing company: {company_name}, index {j}: {str(e)}")
                continue

def main():
    # Define paths to your files
    excel_file_path = pathlib.Path("company_names.xlsx")
    xml_file_path = 'files/.xml'  # Path to the XML file

    # Check if the Excel file exists
    if not excel_file_path.exists():
        print(f"Error: Excel file {excel_file_path} does not exist.")
        sys.exit(1)
    
    # Read the Excel file and process each company name
    try:
        df = pd.read_excel(excel_file_path)
    except Exception as e:
        print(f"Error reading the Excel file: {e}")
        sys.exit(1)

    # Get the company names from the first column and start from the second row
    company_names = df.iloc[:, 0].dropna().tolist()

    # Use ThreadPoolExecutor for parallel processing of company names
    with concurrent.futures.ThreadPoolExecutor() as executor:
        list(tqdm(executor.map(lambda company_name: process_company(company_name, xml_file_path), company_names), desc="Processing company names", total=len(company_names)))

if __name__ == "__main__":
    main()
